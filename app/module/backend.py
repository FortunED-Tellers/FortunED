from openpyxl import load_workbook
from openpyxl import Workbook


def transfer_data(wbkname, pgname, df):
    
    # set file path
    filepath="static/data/admin/" + wbkname
    
     
    workbook_name = filepath
    wb = load_workbook(workbook_name)
    page = wb[pgname]
    page.append(df)
    wb.save(filepath)
    
    # workbook_name = 'KPIs.xlsx'
    # wb = load_workbook(workbook_name)
    # page = wb['YTD_Status']
    # page.append(ytd)
    # wb.save(filename=workbook_name)